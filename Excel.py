import openpyxl

book =openpyxl.load_workbook("C:\\Users\\harih\\Downloads\\workbook(Sheet1).xlsx") # Open the excel
sheet =book.active       # mention active sheet
cell = sheet.cell(row=1,column=3)  # invoking the cell to get the value
print(cell.value)
Dict = {}

sheet.cell(row=2, column=2).value = "Hari" # to enter the data
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)   #  Prints max rows and columns
print(sheet.max_column)
print(sheet['C1'].value)  # Another way to get the value

for i in range(1,sheet.max_row +1): # To get the rows
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column +1):    # TO get the columns
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value # Add to the dict
print(Dict)