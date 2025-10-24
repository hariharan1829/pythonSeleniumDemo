#Download a file update the field in Excel and upload again and check whether the field is updated

import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--start-maximized")

driver = webdriver.Chrome(options=chrome_options)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.implicitly_wait(5)
downloaded_file_path = "C:\\Users\\harih\\Downloads\\download.xlsx"
fruit_name = "Apple"
new_value = "1100"
search_column_fruit = "fruit_name"
search_column_price = "price"
Dict = {}

#download the excel
driver.find_element(By.CSS_SELECTOR,"#downloadButton").click()

#Edit the downloaded excel
book = openpyxl.load_workbook(downloaded_file_path)
sheet = book.active

def update_excel(search_fruit, search_price, updated_value):
    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == search_fruit:
            for j in range (1, sheet.max_row):
                if sheet.cell(row=j, column=i).value == fruit_name:
                    Dict ["Row"] = j
                    print(Dict)

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == search_price:
            Dict["Col"] = i
            print(Dict)

    sheet.cell(row=Dict["Row"], column=Dict["Col"]).value = updated_value
    book.save(downloaded_file_path)

update_excel(search_column_fruit, search_column_price, new_value)

#upload the Downloaded file
   #uploading the file selenium does not automate to go inside the system and upload
   #we find the locator and give the file path in sendkeys and it will work if the locator has file attribute
file_upload = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_upload.send_keys(downloaded_file_path)

#Now wait for the wait for success toast message
wait = WebDriverWait(driver, 5)
toast_locator = (By.XPATH,"//div[@class='Toastify__toast-body']/div[2]")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

#After upload needs to check the updated value in changed in the table
  # im taking example apple and its price with smart xpath
  # will make the same xpath dynamic example if fruit name changes or column of price changes
price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,("//div[text()='"+fruit_name+"']/parent::div/parent::div/div["+price_column+"]")).text

print(actual_price)
assert actual_price == new_value







time.sleep(10)